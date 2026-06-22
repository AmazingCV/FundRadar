from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time
from contextlib import contextmanager
import os
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .report import write_excel, write_markdown
from .utils import ensure_dir, load_yaml, normalize_code, project_path, today_str
from .v4_1.display_utils import deduplicate_fund_display
from .v4_1.signal_aggregator import aggregate_daily_signals


DEFAULT_CATEGORIES = ["全部"]
PROXY_ENV_NAMES = ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]
THEME_PROXY_RULES: list[tuple[str, list[str]]] = [
    ("证券/券商/金融", ["证券", "券商", "金融", "银行", "非银", "保险"]),
    ("煤炭/资源/周期", ["煤炭", "能源", "资源", "有色", "矿业", "周期", "油气", "石油"]),
    ("科技/AI/半导体", ["科技", "半导体", "芯片", "人工智能", "通信", "算力", "电子", "机器人", "软件", "互联网", "计算机"]),
    ("新能源/电力设备", ["新能源", "光伏", "储能", "电池", "电力", "电网", "低碳", "碳中和"]),
    ("医药/创新药", ["医药", "医疗", "创新药", "生物", "健康", "中药"]),
    ("消费", ["消费", "食品", "饮料", "白酒", "家电", "农业"]),
    ("港股/出海", ["港股", "沪港深", "出海", "海外", "全球"]),
]


@dataclass
class IntradayResult:
    as_of: str
    output_dir: Path
    markdown: Path
    excel: Path
    success: bool


def _first_col(columns: list[Any], keywords: list[str]) -> Any | None:
    for col in columns:
        text = str(col)
        if all(k in text for k in keywords):
            return col
    return None


def _to_number(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).replace("%", "").replace(",", "").strip()
    if text in {"", "-", "--", "nan", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_estimation(raw: pd.DataFrame, category: str, fetched_at: str) -> pd.DataFrame:
    if raw is None or raw.empty:
        return pd.DataFrame()
    cols = list(raw.columns)
    code_col = _first_col(cols, ["基金", "代码"]) or _first_col(cols, ["代码"]) or (cols[1] if len(cols) > 1 else None)
    name_col = _first_col(cols, ["基金", "名称"]) or _first_col(cols, ["名称"]) or (cols[2] if len(cols) > 2 else None)
    est_value_col = _first_col(cols, ["估算", "值"])
    est_growth_col = _first_col(cols, ["估算", "增长率"]) or _first_col(cols, ["估算", "涨幅"])
    publish_growth_col = _first_col(cols, ["日增长率"]) or _first_col(cols, ["日涨幅"])
    publish_nav_col = _first_col(cols, ["公布", "单位净值"]) or _first_col(cols, ["单位净值"])
    deviation_col = _first_col(cols, ["估算", "偏差"])

    out = pd.DataFrame()
    out["基金代码"] = raw[code_col].map(normalize_code) if code_col in raw else ""
    out["基金名称"] = raw[name_col].astype(str) if name_col in raw else ""
    out["估算净值"] = raw[est_value_col].map(_to_number) if est_value_col in raw else None
    out["估算涨跌幅%"] = raw[est_growth_col].map(_to_number) if est_growth_col in raw else None
    out["上一公布日涨跌幅%"] = raw[publish_growth_col].map(_to_number) if publish_growth_col in raw else None
    out["上一公布单位净值"] = raw[publish_nav_col].map(_to_number) if publish_nav_col in raw else None
    out["估算偏差"] = raw[deviation_col].map(_to_number) if deviation_col in raw else None
    out["估算分类"] = category
    out["抓取时间"] = fetched_at
    out = out[out["基金代码"].astype(str).str.len().eq(6)].copy()
    return out.drop_duplicates("基金代码").reset_index(drop=True)


def _is_blocking_local_proxy(value: str | None) -> bool:
    return bool(value and "127.0.0.1:9" in str(value))


@contextmanager
def _temporary_disable_blocking_proxy():
    old = {name: os.environ.get(name) for name in PROXY_ENV_NAMES}
    changed = False
    for name, value in old.items():
        if _is_blocking_local_proxy(value):
            os.environ.pop(name, None)
            changed = True
    try:
        yield changed
    finally:
        for name, value in old.items():
            if value is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = value


def fetch_intraday_estimation(categories: list[str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    categories = categories or DEFAULT_CATEGORIES
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows: list[pd.DataFrame] = []
    failures: list[dict[str, Any]] = []
    try:
        import akshare as ak
    except Exception as exc:  # noqa: BLE001
        return pd.DataFrame(), pd.DataFrame([{"分类": "全部", "错误": f"akshare import failed: {exc}", "抓取时间": fetched_at}])

    with _temporary_disable_blocking_proxy() as proxy_disabled:
        for category in categories:
            try:
                raw = ak.fund_value_estimation_em(symbol=category)
                norm = _normalize_estimation(raw, category, fetched_at)
                if proxy_disabled and not norm.empty:
                    norm["代理处理"] = "已临时绕过 127.0.0.1:9 占位代理"
                rows.append(norm)
            except Exception as exc:  # noqa: BLE001 - public intraday endpoint can be unstable.
                failures.append(
                    {
                        "分类": category,
                        "错误": str(exc),
                        "抓取时间": fetched_at,
                        "代理处理": "已临时绕过 127.0.0.1:9 占位代理" if proxy_disabled else "未检测到占位代理",
                    }
                )
    data = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    if not data.empty:
        data = data.drop_duplicates("基金代码").reset_index(drop=True)
    return data, pd.DataFrame(failures)


def _fund_codes_from_df(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["基金代码", "基金名称", "来源"])
    code_col = next((c for c in df.columns if "代码" in str(c)), None)
    name_col = next((c for c in df.columns if "名称" in str(c)), None)
    if code_col is None:
        return pd.DataFrame(columns=["基金代码", "基金名称", "来源"])
    out = pd.DataFrame(
        {
            "基金代码": df[code_col].map(normalize_code),
            "基金名称": df[name_col].astype(str) if name_col is not None else "",
            "来源": source,
        }
    )
    return out.drop_duplicates("基金代码")


def build_watch_universe(as_of: str | None = None) -> pd.DataFrame:
    signals = aggregate_daily_signals(as_of=as_of or today_str(), top_n=50)
    parts = [
        _fund_codes_from_df(signals.selected_pool, "V1精选观察池"),
        _fund_codes_from_df(signals.diversified_pool, "V1分散观察池"),
        _fund_codes_from_df(signals.v3_allocation, "V3组合"),
        _fund_codes_from_df(signals.short_term_top, "V1.1短期异动"),
        _fund_codes_from_df(signals.new_star_top, "V1.1新星"),
    ]
    try:
        watch = pd.DataFrame(load_yaml("config/fund_list.yaml").get("watchlist", []))
        if not watch.empty and "code" in watch.columns:
            parts.append(
                pd.DataFrame(
                    {
                        "基金代码": watch["code"].map(normalize_code),
                        "基金名称": watch.get("name", ""),
                        "来源": "用户观察池",
                    }
                )
            )
    except Exception:
        pass
    out = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=["基金代码", "基金名称", "来源"])
    if out.empty:
        return out
    grouped = out.groupby("基金代码", as_index=False).agg({"基金名称": "first", "来源": lambda x: "、".join(sorted(set(map(str, x))))})
    return grouped.sort_values("来源").reset_index(drop=True)


def _merge_focus(est: pd.DataFrame, universe: pd.DataFrame) -> pd.DataFrame:
    if est.empty or universe.empty:
        return pd.DataFrame()
    out = universe.merge(est, on="基金代码", how="left", suffixes=("_观察池", ""))
    out["基金名称"] = out["基金名称"].where(out["基金名称"].astype(str).str.strip().ne(""), out.get("基金名称_观察池", ""))
    if "基金名称_观察池" in out.columns:
        out = out.drop(columns=["基金名称_观察池"])
    return out.sort_values("估算涨跌幅%", ascending=False, na_position="last").reset_index(drop=True)


def _trade_session_status(now: datetime | None = None) -> str:
    now = now or datetime.now()
    if now.weekday() >= 5:
        return "非交易日，估值可能不更新"
    if time(9, 30) <= now.time() <= time(15, 0):
        return "交易时段内，可用于盘中观察"
    return "非交易时段，估值仅供复盘或接口可用性检查"


def _build_alerts(focus: pd.DataFrame) -> pd.DataFrame:
    if focus.empty or "估算涨跌幅%" not in focus.columns:
        return pd.DataFrame([{"说明": "暂无盘中估值数据，无法生成观察提示"}])
    rows = []
    for _, row in focus.dropna(subset=["估算涨跌幅%"]).iterrows():
        pct = float(row["估算涨跌幅%"])
        if pct >= 3:
            label = "盘中涨幅较高，注意追高和拥挤风险"
        elif pct <= -2:
            label = "盘中下跌较明显，可加入人工观察清单"
        else:
            continue
        rows.append(
            {
                "基金代码": row.get("基金代码", ""),
                "基金名称": row.get("基金名称", ""),
                "来源": row.get("来源", ""),
                "估算涨跌幅%": pct,
                "观察提示": label,
                "口径": "估算净值不是最终净值，不构成买卖建议",
            }
        )
    return pd.DataFrame(rows) if rows else pd.DataFrame([{"说明": "观察池暂无明显盘中异动"}])


def _build_market_temperature(estimation: pd.DataFrame) -> pd.DataFrame:
    if estimation.empty or "估算涨跌幅%" not in estimation.columns:
        return pd.DataFrame([{"项目": "市场温度", "结果": "暂无盘中估值数据", "说明": ""}])
    pct = pd.to_numeric(estimation["估算涨跌幅%"], errors="coerce").dropna()
    if pct.empty:
        return pd.DataFrame([{"项目": "市场温度", "结果": "估算涨跌幅缺失", "说明": ""}])
    up_ratio = float((pct > 0).mean() * 100)
    down_ratio = float((pct < 0).mean() * 100)
    strong_ratio = float((pct >= 2).mean() * 100)
    weak_ratio = float((pct <= -2).mean() * 100)
    median = float(pct.median())
    if median >= 0.5 and up_ratio >= 60:
        label = "偏强"
    elif median <= -0.5 and down_ratio >= 60:
        label = "偏弱"
    else:
        label = "分化/震荡"
    return pd.DataFrame(
        [
            {"项目": "市场温度", "结果": label, "说明": "基于全市场基金盘中估算涨跌幅统计，不代表最终净值"},
            {"项目": "样本数量", "结果": len(pct), "说明": ""},
            {"项目": "平均估算涨跌幅%", "结果": round(float(pct.mean()), 2), "说明": ""},
            {"项目": "中位数估算涨跌幅%", "结果": round(median, 2), "说明": ""},
            {"项目": "上涨基金占比%", "结果": round(up_ratio, 1), "说明": ""},
            {"项目": "下跌基金占比%", "结果": round(down_ratio, 1), "说明": ""},
            {"项目": "估算涨幅>=2%占比%", "结果": round(strong_ratio, 1), "说明": ""},
            {"项目": "估算跌幅<=-2%占比%", "结果": round(weak_ratio, 1), "说明": ""},
            {"项目": "最高估算涨跌幅%", "结果": round(float(pct.max()), 2), "说明": ""},
            {"项目": "最低估算涨跌幅%", "结果": round(float(pct.min()), 2), "说明": ""},
        ]
    )


def _representative_funds(df: pd.DataFrame, limit: int = 5) -> str:
    if df.empty:
        return ""
    shown = deduplicate_fund_display(df.head(max(limit * 3, limit))).head(limit)
    items: list[str] = []
    for _, row in shown.iterrows():
        code = normalize_code(row.get("基金代码", ""))
        name = str(row.get("基金名称", "")).strip()
        pct = _to_number(row.get("估算涨跌幅%"))
        if pct is None:
            items.append(f"{code} {name}")
        else:
            items.append(f"{code} {name}({pct:+.2f}%)")
    return "；".join(items)


def _build_theme_proxy(estimation: pd.DataFrame) -> pd.DataFrame:
    if estimation.empty or "基金名称" not in estimation.columns or "估算涨跌幅%" not in estimation.columns:
        return pd.DataFrame([{"说明": "暂无盘中估值数据，无法生成主题代理热度"}])
    rows: list[dict[str, Any]] = []
    for theme, keywords in THEME_PROXY_RULES:
        pattern = "|".join(keywords)
        sub = estimation[estimation["基金名称"].astype(str).str.contains(pattern, regex=True, na=False)].copy()
        sub["估算涨跌幅%"] = pd.to_numeric(sub["估算涨跌幅%"], errors="coerce")
        sub = sub.dropna(subset=["估算涨跌幅%"])
        if sub.empty:
            continue
        top = sub.sort_values("估算涨跌幅%", ascending=False)
        rows.append(
            {
                "主题代理": theme,
                "匹配基金数": len(sub),
                "平均估算涨跌幅%": round(float(sub["估算涨跌幅%"].mean()), 2),
                "中位数估算涨跌幅%": round(float(sub["估算涨跌幅%"].median()), 2),
                "上涨占比%": round(float((sub["估算涨跌幅%"] > 0).mean() * 100), 1),
                "强势基金数(>=2%)": int((sub["估算涨跌幅%"] >= 2).sum()),
                "弱势基金数(<=-1%)": int((sub["估算涨跌幅%"] <= -1).sum()),
                "代表基金": _representative_funds(top, limit=5),
                "口径": "基金名称关键词代理，不等同真实持仓主题或行业资金流",
            }
        )
    if not rows:
        return pd.DataFrame([{"说明": "未匹配到可用主题代理"}])
    return pd.DataFrame(rows).sort_values(["平均估算涨跌幅%", "上涨占比%"], ascending=False).reset_index(drop=True)


def _build_focus_source_summary(focus: pd.DataFrame) -> pd.DataFrame:
    if focus.empty or "来源" not in focus.columns or "估算涨跌幅%" not in focus.columns:
        return pd.DataFrame([{"说明": "暂无观察池盘中估值数据"}])
    rows: list[dict[str, Any]] = []
    for source in ["V1精选观察池", "V1分散观察池", "V3组合", "V1.1短期异动", "V1.1新星", "用户观察池"]:
        sub = focus[focus["来源"].astype(str).str.contains(source, regex=False, na=False)].copy()
        sub["估算涨跌幅%"] = pd.to_numeric(sub["估算涨跌幅%"], errors="coerce")
        sub = sub.dropna(subset=["估算涨跌幅%"])
        if sub.empty:
            continue
        rows.append(
            {
                "观察范围": source,
                "匹配基金数": len(sub),
                "平均估算涨跌幅%": round(float(sub["估算涨跌幅%"].mean()), 2),
                "中位数估算涨跌幅%": round(float(sub["估算涨跌幅%"].median()), 2),
                "上涨占比%": round(float((sub["估算涨跌幅%"] > 0).mean() * 100), 1),
                "涨幅靠前": _representative_funds(sub.sort_values("估算涨跌幅%", ascending=False), limit=3),
                "跌幅靠前": _representative_funds(sub.sort_values("估算涨跌幅%", ascending=True), limit=3),
            }
        )
    return pd.DataFrame(rows) if rows else pd.DataFrame([{"说明": "暂无观察池盘中估值数据"}])


def _build_intraday_brief(
    market_temperature: pd.DataFrame,
    theme_proxy: pd.DataFrame,
    focus: pd.DataFrame,
    failures: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if not market_temperature.empty:
        first = market_temperature.iloc[0]
        rows.append({"类型": "市场温度", "摘要": f"{first.get('结果', '')}", "口径": first.get("说明", "")})
    if not theme_proxy.empty and "主题代理" in theme_proxy.columns:
        top_theme = theme_proxy.iloc[0]
        rows.append(
            {
                "类型": "盘中较强方向",
                "摘要": f"{top_theme.get('主题代理', '')}，平均估算涨跌幅 {top_theme.get('平均估算涨跌幅%', '')}%",
                "口径": "基金名称关键词代理，不等同真实行业资金流",
            }
        )
        weak = theme_proxy.sort_values("平均估算涨跌幅%", ascending=True).iloc[0]
        rows.append(
            {
                "类型": "盘中较弱方向",
                "摘要": f"{weak.get('主题代理', '')}，平均估算涨跌幅 {weak.get('平均估算涨跌幅%', '')}%",
                "口径": "基金名称关键词代理",
            }
        )
    valid_focus = focus.copy()
    if not valid_focus.empty:
        valid_focus["估算涨跌幅%"] = pd.to_numeric(valid_focus.get("估算涨跌幅%"), errors="coerce")
        valid_focus = valid_focus.dropna(subset=["估算涨跌幅%"])
    if not valid_focus.empty:
        rows.append(
            {
                "类型": "观察池涨幅靠前",
                "摘要": _representative_funds(valid_focus.sort_values("估算涨跌幅%", ascending=False), limit=5),
                "口径": "来自 V1/V1.1/V3/用户观察池交集",
            }
        )
        rows.append(
            {
                "类型": "观察池跌幅靠前",
                "摘要": _representative_funds(valid_focus.sort_values("估算涨跌幅%", ascending=True), limit=5),
                "口径": "来自 V1/V1.1/V3/用户观察池交集",
            }
        )
    rows.append(
        {
            "类型": "接口状态",
            "摘要": "正常" if failures.empty else f"{len(failures)} 个分类抓取失败",
            "口径": "失败分类不影响已抓取数据",
        }
    )
    return pd.DataFrame(rows)


def _format_code_columns_as_text(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        header = [cell.value for cell in ws[1]]
        for idx, value in enumerate(header, start=1):
            if value and "代码" in str(value):
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                    for item in cell:
                        if item.value is not None and str(item.value).strip():
                            item.value = normalize_code(item.value)
                            item.number_format = "@"
    wb.save(path)


def run_intraday_estimation(
    categories: list[str] | None = None,
    as_of: str | None = None,
    top_n: int = 30,
    output_dir: str | Path | None = None,
) -> IntradayResult:
    date_text = as_of or today_str()
    out_dir = ensure_dir(output_dir or project_path("reports", "intraday", date_text))
    estimation, failures = fetch_intraday_estimation(categories)
    universe = build_watch_universe(date_text)
    focus = _merge_focus(estimation, universe)
    alerts = _build_alerts(focus)
    market_temperature = _build_market_temperature(estimation)
    theme_proxy = _build_theme_proxy(estimation)
    focus_source_summary = _build_focus_source_summary(focus)
    intraday_brief = _build_intraday_brief(market_temperature, theme_proxy, focus, failures)
    focus_display = deduplicate_fund_display(focus).head(top_n) if not focus.empty else pd.DataFrame()

    if not estimation.empty:
        top_up = estimation.sort_values("估算涨跌幅%", ascending=False, na_position="last").head(top_n).reset_index(drop=True)
        top_down = estimation.sort_values("估算涨跌幅%", ascending=True, na_position="last").head(top_n).reset_index(drop=True)
    else:
        top_up = pd.DataFrame()
        top_down = pd.DataFrame()

    summary = pd.DataFrame(
        [
            {"项目": "报告日期", "结果": date_text, "说明": ""},
            {"项目": "交易时段状态", "结果": _trade_session_status(), "说明": ""},
            {"项目": "估算样本数", "结果": len(estimation), "说明": "来自 AKShare 东方财富净值估算接口"},
            {"项目": "观察池匹配数", "结果": int(focus["估算涨跌幅%"].notna().sum()) if not focus.empty else 0, "说明": "匹配 V1/V1.1/V3/用户观察池"},
            {"项目": "接口失败数", "结果": len(failures), "说明": "失败不影响其他模块"},
            {"项目": "核心口径", "结果": "估算净值不是最终净值", "说明": "只用于下午3点前的人工观察，不构成买卖建议"},
        ]
    )

    sheets = {
        "盘中摘要": summary,
        "盘中变化摘要": intraday_brief,
        "市场温度": market_temperature,
        "主题代理热度": theme_proxy,
        "观察池来源表现": focus_source_summary,
        "观察池盘中估值": focus,
        "盘中观察提示": alerts,
        "估算涨幅Top": top_up,
        "估算跌幅Top": top_down,
        "估值总表": estimation,
        "观察池代码来源": universe,
        "接口失败记录": failures if not failures.empty else pd.DataFrame([{"说明": "接口正常或未发现失败"}]),
        "口径说明": pd.DataFrame(
            [
                {"说明": "盘中净值估算来自公开估值接口，可能与最终公布净值有偏差。"},
                {"说明": "本模块不修改 V1/V1.1/V2-lite/V3/V4 任何评分、仓位、主题或风控逻辑。"},
                {"说明": "盘中提示只用于人工观察，不是买卖建议。"},
            ]
        ),
    }
    excel = write_excel(out_dir / "intraday_estimation.xlsx", sheets)
    _format_code_columns_as_text(excel)
    markdown = write_markdown(
        out_dir / "intraday_estimation.md",
        f"FundRadar 盘中净值估算观察 {date_text}",
        {
            "定位": "只读盘中观察层。估算净值不是最终净值，不预测收益，不提供买卖建议。",
            "盘中摘要": summary,
            "盘中变化摘要": intraday_brief,
            "市场温度": market_temperature,
            "主题代理热度": theme_proxy.head(10) if "主题代理" in theme_proxy.columns else theme_proxy,
            "观察池来源表现": focus_source_summary,
            "观察池盘中估值Top": focus_display if not focus_display.empty else pd.DataFrame([{"说明": "暂无观察池估值数据"}]),
            "盘中观察提示": alerts,
            "接口失败记录": failures if not failures.empty else pd.DataFrame([{"说明": "接口正常或未发现失败"}]),
        },
    )
    return IntradayResult(date_text, out_dir, markdown, excel, success=not estimation.empty)
